# hr/services.py
import os
import io
import zipfile
import tempfile
from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.db.models import Avg, Count
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage

from users.models import CustomUser, Department, StaffProfile, StaffAppraisal, PerformanceTarget
from spe.models import SPEPeriod, SelfAssessment, SupervisorEvaluation as SpeSupervisorEvaluation
from spe.models import TeachingStaffEvaluation, NonTeachingStaffEvaluation
from .models import SupervisorAppraisal


class BulkReportService:
    """Service for generating bulk department reports"""
    
    @staticmethod
    def get_department_evaluated_staff(department_id, period_id=None):
        """Get all evaluated staff in a department for a given period"""
        department = get_object_or_404(Department, id=department_id)
        
        # Get active period if not specified
        if not period_id:
            period = SPEPeriod.objects.filter(is_active=True).first()
        else:
            period = get_object_or_404(SPEPeriod, id=period_id)
        
        if not period:
            return None, None, []
        
        # Get all evaluated staff
        evaluated_staff = []
        
        # Regular staff with StaffAppraisal
        regular_appraisals = StaffAppraisal.objects.filter(
            profile__user__department=department,
            period=period,
            overall_score__isnull=False
        ).select_related('profile__user', 'period')
        
        for appraisal in regular_appraisals:
            evaluated_staff.append({
                'type': 'regular',
                'appraisal': appraisal,
                'user': appraisal.profile.user,
                'score': appraisal.overall_score,
                'status': appraisal.status,
                'date': appraisal.updated_at
            })
        
        # Supervisors with SupervisorAppraisal
        supervisor_appraisals = SupervisorAppraisal.objects.filter(
            supervisor__department=department,
            period=period,
            overall_score__isnull=False
        ).select_related('supervisor', 'period')
        
        for appraisal in supervisor_appraisals:
            evaluated_staff.append({
                'type': 'supervisor',
                'appraisal': appraisal,
                'user': appraisal.supervisor,
                'score': appraisal.overall_score,
                'status': appraisal.status,
                'date': appraisal.evaluated_at
            })
        
        return department, period, evaluated_staff
    
    @staticmethod
    def get_performance_level(score):
        """Get performance level based on score"""
        if score >= 90:
            return "Outstanding", "#10b981"
        elif score >= 80:
            return "Exceeds Expectations", "#3b82f6"
        elif score >= 50:
            return "Meets Expectations", "#f59e0b"
        elif score >= 30:
            return "Below Expectations", "#f97316"
        else:
            return "Far Below Expectations", "#ef4444"
    
    @staticmethod
    def generate_department_excel_report(department, period, evaluated_staff):
        """Generate Excel report for department"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Department Report"
        
        # Styles
        title_font = Font(bold=True, size=16, color="2c5aa0")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2c5aa0", end_color="2c5aa0", fill_type="solid")
        even_row_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Title and headers
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{department.name} - Performance Report"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Period: {period.name} | Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].alignment = center_align
        
        # Headers
        headers = ["#", "Staff Name", "Staff ID", "Department", "Role", 
                  "Performance Score", "Rating", "Status", "Date"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        # Data rows
        for i, staff_data in enumerate(evaluated_staff, start=1):
            row = i + 4
            user = staff_data['user']
            score = staff_data['score']
            perf_level, perf_color = BulkReportService.get_performance_level(score)
            
            data = [
                i,
                user.get_full_name(),
                user.username or user.email,
                department.name,
                user.get_role_display(),
                f"{score:.1f}%",
                perf_level,
                staff_data['status'].title(),
                staff_data['date'].strftime('%Y-%m-%d') if staff_data['date'] else "N/A"
            ]
            
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                
                if col in [1, 6, 8]:  # Number, Score, Status columns
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                
                # Alternate row colors
                if i % 2 == 0:
                    cell.fill = even_row_fill
                else:
                    cell.fill = odd_row_fill
                
                # Color code performance rating
                if col == 7:  # Rating column
                    cell.font = Font(bold=True, color=perf_color[1:])  # Remove #
                
                # Color code scores
                if col == 6:  # Score column
                    if score >= 90:
                        cell.font = Font(bold=True, color="10b981")
                    elif score >= 80:
                        cell.font = Font(bold=True, color="3b82f6")
                    elif score >= 50:
                        cell.font = Font(bold=True, color="f59e0b")
                    elif score >= 30:
                        cell.font = Font(bold=True, color="f97316")
                    else:
                        cell.font = Font(bold=True, color="ef4444")
        
        # Summary statistics
        summary_row = len(evaluated_staff) + 6
        avg_score = sum(s['score'] for s in evaluated_staff) / len(evaluated_staff) if evaluated_staff else 0
        
        summary_data = [
            ["Department:", department.name],
            ["Total Evaluated Staff:", len(evaluated_staff)],
            ["Average Score:", f"{avg_score:.1f}%"],
            ["Report Period:", period.name],
            ["Generated On:", timezone.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for i, (label, value) in enumerate(summary_data, start=summary_row):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
        
        # Performance scale legend
        legend_row = summary_row + len(summary_data) + 2
        ws.cell(row=legend_row, column=1, value="Performance Scale:").font = Font(bold=True)
        
        scale_data = [
            ("Outstanding", "90-100"),
            ("Exceeds Expectations", "80-89"),
            ("Meets Expectations", "50-79"),
            ("Below Expectations", "30-49"),
            ("Far Below Expectations", "Below 30")
        ]
        
        for i, (level, range_) in enumerate(scale_data, start=legend_row + 1):
            ws.cell(row=i, column=1, value=level)
            ws.cell(row=i, column=2, value=range_)
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{department.name.replace(' ', '_')}_Performance_Report_{period.name.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response
    
    @staticmethod
    def generate_department_pdf_report(department, period, evaluated_staff):
        """Generate PDF report for department"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=landscape(A4),
            topMargin=0.5*inch,
            bottomMargin=0.5*inch,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch
        )
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,
            textColor=colors.HexColor("#2c5aa0")
        )
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=15,
            alignment=1,
            textColor=colors.HexColor("#666666")
        )
        
        # Title
        story.append(Paragraph(f"{department.name} - Performance Report", title_style))
        story.append(Paragraph(f"Appraisal Period: {period.name}", subtitle_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Summary statistics
        total_staff = len(evaluated_staff)
        avg_score = sum(s['score'] for s in evaluated_staff) / total_staff if total_staff > 0 else 0
        
        summary_data = [
            ["Total Evaluated Staff:", str(total_staff)],
            ["Average Score:", f"{avg_score:.1f}%"],
            ["Period:", period.name],
            ["Generated:", timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Department Head:", department.head.get_full_name() if department.head else "Not Assigned"]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2c5aa0")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Staff performance table
        if evaluated_staff:
            story.append(Paragraph("Staff Performance Summary", styles['Heading2']))
            
            table_data = []
            headers = ["#", "Staff Name", "Role", "Score", "Performance Level", "Status"]
            table_data.append(headers)
            
            for i, staff_data in enumerate(evaluated_staff, start=1):
                user = staff_data['user']
                score = staff_data['score']
                perf_level, _ = BulkReportService.get_performance_level(score)
                
                row = [
                    str(i),
                    user.get_full_name(),
                    user.get_role_display(),
                    f"{score:.1f}%",
                    perf_level,
                    staff_data['status'].title()
                ]
                table_data.append(row)
            
            table = Table(table_data, colWidths=[0.5*inch, 2*inch, 1.5*inch, 1*inch, 2*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2c5aa0")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
            ]))
            story.append(table)
        
        # Performance scale legend
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph("Performance Rating Scale", styles['Heading3']))
        
        scale_data = [
            ["Rating", "Score Range", "Description"],
            ["Outstanding", "90-100", "Exceptional performance exceeding all expectations"],
            ["Exceeds Expectations", "80-89", "Performance consistently above required standards"],
            ["Meets Expectations", "50-79", "Performance meeting all required standards"],
            ["Below Expectations", "30-49", "Performance below required standards, improvement needed"],
            ["Far Below Expectations", "Below 30", "Performance significantly below expectations"],
        ]
        
        scale_table = Table(scale_data, colWidths=[1.5*inch, 1*inch, 3*inch])
        scale_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#6c757d")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 1), (0, 1), colors.HexColor("#d1fae5")),
            ('BACKGROUND', (0, 2), (0, 2), colors.HexColor("#dbeafe")),
            ('BACKGROUND', (0, 3), (0, 3), colors.HexColor("#fef3c7")),
            ('BACKGROUND', (0, 4), (0, 4), colors.HexColor("#fed7aa")),
            ('BACKGROUND', (0, 5), (0, 5), colors.HexColor("#fecaca")),
        ]))
        story.append(scale_table)
        
        # Footer
        story.append(Spacer(1, 0.3*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            textColor=colors.HexColor("#666666"),
        )
        footer = Paragraph(
            f"Generated by HR System | {timezone.now().strftime('%Y-%m-d at %H:%M')}",
            footer_style,
        )
        story.append(footer)
        
        # Build PDF
        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()
        
        # Create response
        response = HttpResponse(content_type='application/pdf')
        filename = f"{department.name.replace(' ', '_')}_Report_{period.name.replace(' ', '_')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        
        return response
    
    @staticmethod
    def generate_individual_reports_zip(request, department, period, evaluated_staff):
        """Generate ZIP file with individual reports"""
        import tempfile
        import zipfile
        
        temp_dir = tempfile.mkdtemp()
        zip_filename = f"{department.name.replace(' ', '_')}_Individual_Reports_{period.name.replace(' ', '_')}.zip"
        
        try:
            generated_count = 0
            
            # For each staff member, generate individual report
            for staff_data in evaluated_staff:
                try:
                    if staff_data['type'] == 'regular':
                        appraisal_id = staff_data['appraisal'].id
                        
                        # Use IndividualReportService
                        eval_data = IndividualReportService.get_staff_evaluation_data(appraisal_id)
                        
                        # Validate data
                        is_valid, missing_items = IndividualReportService.validate_evaluation_data(eval_data)
                        
                        if is_valid:
                            # Generate PDF
                            pdf_content = IndividualReportService.generate_evaluation_pdf(eval_data)
                            
                            filename = f"{staff_data['user'].get_full_name().replace(' ', '_')}_Report.pdf"
                            filepath = os.path.join(temp_dir, filename)
                            
                            with open(filepath, 'wb') as f:
                                f.write(pdf_content)
                            generated_count += 1
                        else:
                            print(f"Skipping {staff_data['user'].get_full_name()}: Missing {missing_items}")
                            continue
                except Exception as e:
                    print(f"Error generating PDF for {staff_data['user'].get_full_name()}: {e}")
                    continue
            
            if generated_count == 0:
                # Clean up
                for file in os.listdir(temp_dir):
                    os.remove(os.path.join(temp_dir, file))
                os.rmdir(temp_dir)
                return None
            
            # Create ZIP file
            zip_path = os.path.join(temp_dir, zip_filename)
            with zipfile.ZipFile(zip_path, 'w') as zipf:
                for file in os.listdir(temp_dir):
                    if file.endswith('.pdf'):
                        zipf.write(os.path.join(temp_dir, file), file)
            
            # Read ZIP file
            with open(zip_path, 'rb') as f:
                zip_data = f.read()
            
            # Clean up
            for file in os.listdir(temp_dir):
                os.remove(os.path.join(temp_dir, file))
            os.rmdir(temp_dir)
            
            return zip_data, zip_filename, generated_count
            
        except Exception as e:
            # Clean up on error
            try:
                for file in os.listdir(temp_dir):
                    os.remove(os.path.join(temp_dir, file))
                os.rmdir(temp_dir)
            except:
                pass
            raise e
    
    @staticmethod
    def get_department_performance_summary(department, period, evaluated_staff):
        """Get department performance summary statistics"""
        total_staff = len(evaluated_staff)
        scores = [staff['score'] for staff in evaluated_staff]
        
        if total_staff > 0:
            avg_score = sum(scores) / total_staff
            max_score = max(scores)
            min_score = min(scores)
            
            # Count by performance level
            performance_levels = {
                'Outstanding': 0,
                'Exceeds Expectations': 0,
                'Meets Expectations': 0,
                'Below Expectations': 0,
                'Far Below Expectations': 0
            }
            
            for score in scores:
                level, _ = BulkReportService.get_performance_level(score)
                performance_levels[level] += 1
            
            return {
                'total_staff': total_staff,
                'avg_score': round(avg_score, 2),
                'max_score': round(max_score, 2),
                'min_score': round(min_score, 2),
                'performance_levels': performance_levels,
                'department': department,
                'period': period,
                'evaluated_staff': evaluated_staff
            }
        return None
    
    @staticmethod
    def generate_department_summary_excel(department, period, evaluated_staff):
        """Generate Excel with department summary statistics"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Department Summary"
        
        # Add summary statistics
        ws.append(["Department Performance Summary"])
        ws.append([f"Department: {department.name}"])
        ws.append([f"Period: {period.name}"])
        ws.append([f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"])
        ws.append([])
        
        # Statistics table
        total_staff = len(evaluated_staff)
        scores = [s['score'] for s in evaluated_staff]
        avg_score = sum(scores) / total_staff if total_staff > 0 else 0
        
        stats_data = [
            ["Total Evaluated Staff", total_staff],
            ["Average Score", f"{avg_score:.1f}%"],
            ["Highest Score", f"{max(scores):.1f}%" if scores else "0%"],
            ["Lowest Score", f"{min(scores):.1f}%" if scores else "0%"],
        ]
        
        for row in stats_data:
            ws.append(row)
        
        ws.append([])
        ws.append(["Performance Level Distribution"])
        
        # Performance level counts
        level_counts = {}
        for staff in evaluated_staff:
            level, _ = BulkReportService.get_performance_level(staff['score'])
            level_counts[level] = level_counts.get(level, 0) + 1
        
        for level, count in level_counts.items():
            ws.append([level, count])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{department.name.replace(' ', '_')}_Summary_{period.name.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class IndividualReportService:
    """Service for generating individual staff evaluation reports"""
    
    @staticmethod
    def get_staff_evaluation_data(appraisal_id):
        """Get all evaluation data for a staff member"""
        appraisal = get_object_or_404(
            StaffAppraisal.objects.select_related(
                "profile__user", "period", "profile__user__department"
            ),
            id=appraisal_id,
        )
        
        staff_user = appraisal.profile.user
        period = appraisal.period
        
        # Get all evaluation data
        data = {
            'appraisal': appraisal,
            'staff_user': staff_user,
            'period': period,
            'performance_targets': PerformanceTarget.objects.filter(
                staff=staff_user, period=period
            ),
            'self_assessments': SelfAssessment.objects.filter(
                staff=staff_user, period=period
            ),
            'supervisor_evaluations': SpeSupervisorEvaluation.objects.filter(
                supervisor__department=staff_user.department, period=period
            ),
        }
        
        # Add formal evaluations based on role
        if staff_user.role == "teaching":
            data['formal_evaluations'] = TeachingStaffEvaluation.objects.filter(
                staff=staff_user, period=period
            )
        else:
            data['formal_evaluations'] = NonTeachingStaffEvaluation.objects.filter(
                staff=staff_user, period=period
            )
        
        return data
    
    @staticmethod
    def validate_evaluation_data(data):
        """Validate that staff has required evaluation data"""
        has_both_evaluations = (
            data['performance_targets'].exists() and 
            data['self_assessments'].exists()
        )
        
        if not has_both_evaluations:
            missing_items = []
            if not data['performance_targets'].exists():
                missing_items.append("performance targets")
            if not data['self_assessments'].exists():
                missing_items.append("self-assessments")
            return False, missing_items
        
        return True, []
    
    @staticmethod
    def generate_evaluation_pdf(data):
        """Generate PDF report for staff evaluation"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
        )
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=14,
            spaceAfter=20,
            alignment=1,
            textColor=colors.HexColor("#2c5aa0"),
        )
        title = Paragraph("STAFF EVALUATION REPORT", title_style)
        story.append(title)
        
        # Subtitle
        subtitle_style = ParagraphStyle(
            "Subtitle",
            parent=styles["Normal"],
            fontSize=12,
            spaceAfter=20,
            alignment=1,
            textColor=colors.HexColor("#666666"),
        )
        subtitle = Paragraph(
            f"{data['staff_user'].get_full_name()} - {data['period'].name}", 
            subtitle_style
        )
        story.append(subtitle)
        story.append(Spacer(1, 0.2 * inch))
        
        # Basic Information
        story.append(Paragraph("Basic Information", styles["Heading2"]))
        
        info_data = [
            ["Staff Name:", data['staff_user'].get_full_name()],
            [
                "Department:",
                data['staff_user'].department.name if data['staff_user'].department else "N/A",
            ],
            ["Role:", data['staff_user'].get_role_display()],
            ["Period:", data['period'].name],
            [
                "Overall Score:",
                (
                    f"{data['appraisal'].overall_score}%"
                    if data['appraisal'].overall_score
                    else "Not Scored"
                ),
            ],
        ]
        
        info_table = Table(info_data, colWidths=[1.5 * inch, 3 * inch])
        info_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(info_table)
        story.append(Spacer(1, 0.3 * inch))
        
        # Performance Targets Section
        if data['performance_targets'].exists():
            story.append(Paragraph("Performance Targets", styles["Heading2"]))
            
            target_data = [["#", "Description", "Status", "Rating"]]
            for target in data['performance_targets']:
                target_data.append(
                    [
                        str(target.target_number),
                        Paragraph(
                            (
                                target.description[:40] + "..."
                                if len(target.description) > 40
                                else target.description
                            ),
                            styles["Normal"],
                        ),
                        target.get_status_display(),
                        (
                            f"{target.performance_rating}%"
                            if target.performance_rating
                            else "N/A"
                        ),
                    ]
                )
            
            target_table = Table(
                target_data,
                colWidths=[0.4 * inch, 2.5 * inch, 1 * inch, 0.8 * inch],
            )
            target_table.setStyle(
                TableStyle(
                    [
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, 0),
                            colors.HexColor("#2c5aa0"),
                        ),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        (
                            "ALIGN",
                            (3, 1),
                            (3, -1),
                            "CENTER",
                        ),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ]
                )
            )
            story.append(target_table)
            story.append(Spacer(1, 0.2 * inch))
        
        # Self-Assessments Section
        if data['self_assessments'].exists():
            story.append(Paragraph("Self-Assessments", styles["Heading2"]))
            
            self_data = [["Attribute", "Rating", "Indicator"]]
            for assessment in data['self_assessments']:
                self_data.append(
                    [
                        assessment.attribute.name,
                        f"{assessment.self_rating}/5",
                        Paragraph(
                            (
                                assessment.indicator.description[:50] + "..."
                                if len(assessment.indicator.description) > 50
                                else assessment.indicator.description
                            ),
                            styles["Normal"],
                        ),
                    ]
                )
            
            self_table = Table(
                self_data, colWidths=[1.5 * inch, 0.6 * inch, 2.6 * inch]
            )
            self_table.setStyle(
                TableStyle(
                    [
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, 0),
                            colors.HexColor("#28a745"),
                        ),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        (
                            "ALIGN",
                            (1, 1),
                            (1, -1),
                            "CENTER",
                        ),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ]
                )
            )
            story.append(self_table)
            story.append(Spacer(1, 0.2 * inch))
        
        # Supervisor Evaluations Section
        if data['supervisor_evaluations'].exists():
            story.append(Paragraph("Supervisor Evaluations", styles["Heading2"]))
            
            supervisor_data = [["Attribute", "Rating"]]
            for evaluation in data['supervisor_evaluations']:
                supervisor_data.append(
                    [evaluation.attribute.name, f"{evaluation.rating}/5"]
                )
            
            supervisor_table = Table(
                supervisor_data, colWidths=[3 * inch, 0.8 * inch]
            )
            supervisor_table.setStyle(
                TableStyle(
                    [
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, 0),
                            colors.HexColor("#dc3545"),
                        ),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        (
                            "ALIGN",
                            (1, 1),
                            (1, -1),
                            "CENTER",
                        ),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ]
                )
            )
            story.append(supervisor_table)
            story.append(Spacer(1, 0.2 * inch))
        
        # Summary Statistics
        story.append(Paragraph("Summary", styles["Heading2"]))
        
        summary_data = [
            ["Evaluation Type", "Count", "Avg Rating"],
            [
                "Targets",
                str(data['performance_targets'].count()),
                f"{data['performance_targets'].aggregate(avg=Avg('performance_rating'))['avg'] or 0:.1f}%",
            ],
            [
                "Self-Assessments",
                str(data['self_assessments'].count()),
                f"{data['self_assessments'].aggregate(avg=Avg('self_rating'))['avg'] or 0:.1f}/5",
            ],
            [
                "Supervisor Evals",
                str(data['supervisor_evaluations'].count()),
                f"{data['supervisor_evaluations'].aggregate(avg=Avg('rating'))['avg'] or 0:.1f}/5",
            ],
        ]
        
        summary_table = Table(
            summary_data, colWidths=[2 * inch, 1 * inch, 1 * inch]
        )
        summary_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#6c757d"),
                    ),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    (
                        "ALIGN",
                        (1, 1),
                        (2, -1),
                        "CENTER",
                    ),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )
        story.append(summary_table)
        
        # Footer
        story.append(Spacer(1, 0.3 * inch))
        footer_style = ParagraphStyle(
            "Footer",
            parent=styles["Normal"],
            fontSize=8,
            alignment=1,
            textColor=colors.HexColor("#666666"),
        )
        footer = Paragraph(
            f"Generated on {timezone.now().strftime('%Y-%m-%d at %H:%M')} - KyU HR System",
            footer_style,
        )
        story.append(footer)
        
        # Build PDF
        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()
        
        return pdf
    
    @staticmethod
    def generate_evaluation_excel(data):
        """Generate Excel report for staff evaluation"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Evaluation Report"
        
        # Styles
        title_font = Font(bold=True, size=14, color="2c5aa0")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2c5aa0", end_color="2c5aa0", fill_type="solid")
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = f"STAFF EVALUATION REPORT - {data['staff_user'].get_full_name()}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Basic Information
        ws['A3'] = "Basic Information"
        ws['A3'].font = Font(bold=True)
        
        basic_info = [
            ["Staff Name:", data['staff_user'].get_full_name()],
            ["Department:", data['staff_user'].department.name if data['staff_user'].department else "N/A"],
            ["Role:", data['staff_user'].get_role_display()],
            ["Period:", data['period'].name],
            ["Overall Score:", f"{data['appraisal'].overall_score}%" if data['appraisal'].overall_score else "Not Scored"],
        ]
        
        for i, (label, value) in enumerate(basic_info, start=4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
        
        # Performance Targets
        if data['performance_targets'].exists():
            start_row = len(basic_info) + 6
            ws[f'A{start_row}'] = "Performance Targets"
            ws[f'A{start_row}'].font = Font(bold=True)
            
            headers = ["#", "Description", "Status", "Rating"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row+1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
            
            for i, target in enumerate(data['performance_targets'], start=1):
                row = start_row + 1 + i
                ws.cell(row=row, column=1, value=target.target_number)
                ws.cell(row=row, column=2, value=target.description[:100])
                ws.cell(row=row, column=3, value=target.get_status_display())
                ws.cell(row=row, column=4, value=target.performance_rating or "N/A")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        return wb
        